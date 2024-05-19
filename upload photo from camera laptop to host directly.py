import cv2
import paramiko
from openpyxl import Workbook, load_workbook
import threading

# Function to capture photo from a specific camera
def capture_photo(photo_name, camera_id=1):
    cap = cv2.VideoCapture(camera_id)
    if not cap.isOpened():
        print(f"Error: Could not open camera with ID {camera_id}.")
        return None

    ret, frame = cap.read()
    if not ret:
        print("Error: Could not read frame from camera.")
        return None

    photo_path = f"{photo_name}.jpg"
    cv2.imwrite(photo_path, frame)
    cap.release()
    return photo_path

# Function to upload photo via SFTP
def upload_photo(local_path, remote_path, hostname, port, username, password, lock, ws, photo_name, photo_url):
    try:
        transport = paramiko.Transport((hostname, port))
        transport.connect(username=username, password=password)
        
        sftp = paramiko.SFTPClient.from_transport(transport)
        sftp.put(local_path, remote_path)
        
        sftp.close()
        transport.close()
        
        # Add data to Excel workbook
        with lock:
            row = [photo_name, photo_url]
            ws.append(row)
        
        print(f"Photo {photo_name} captured and uploaded.")
        return True
    except Exception as e:
        print(f"Error: {e}")
        return False

if __name__ == "__main__":
    # Check if the Excel file already exists
    try:
        wb = load_workbook("photo_links.xlsx")
        ws = wb.active
    except FileNotFoundError:
        # Create a new Excel workbook if it doesn't exist
        wb = Workbook()
        ws = wb.active
        ws.append(["Image Name", "Image URL"])

    hostname = "hostname"
    port = 22
    username = "username"
    password = "password"

    lock = threading.Lock()

    while True:
        # Enter the number of photos you want to take
        num_photos = int(input("Enter the number of photos you want to take: "))

        threads = []

        for i in range(num_photos):
            # Enter the unique photo name
            photo_name = input(f"Enter the name for photo {i+1}: ")
            photo_path = capture_photo(photo_name)

            if photo_path:
                remote_photo_path = f"/home/username/html/wp-content/uploads/new/{photo_name}.jpg"
                photo_url = f"https://domain.com/wp-content/uploads/new/{photo_name}.jpg"

                # Start a new thread for uploading the photo
                thread = threading.Thread(target=upload_photo, args=(photo_path, remote_photo_path, hostname, port, username, password, lock, ws, photo_name, photo_url))
                thread.start()
                threads.append(thread)

        # Wait for all threads to complete
        for thread in threads:
            thread.join()

        # Save the Excel file
        excel_file = "photo_links.xlsx"
        wb.save(excel_file)
        print(f"Excel file '{excel_file}' updated with new photo links.")

        # Ask if the user wants to continue
        continue_input = input("Do you want to take more photos for more products? (Y/N): ")
        if continue_input.lower() != 'y':
            break
